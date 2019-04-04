#  coding utf-8
# @time      :2019/3/1111:53
# @Author    :zjunbin
# @Email     :648060307@qq.com
# @File      :read_excel.py
from openpyxl import load_workbook
from common.mylog import MyLog
from common import constants
data_file = constants.data_case
mylog = MyLog()

class Case:
    def __init__(self):
        self.caseid = None
        self.title = None
        self.url = None
        self.meathod = None
        self.params = None
        self.expected = None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#Excel工作簿文件名或地址
        self.sheet_name=sheet_name#表单名

    def get_case(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        max_row = sheet.max_row
        data_list = []
        for r in range(2, max_row + 1):
            case = Case()
            case.caseid = sheet.cell(r, 1).value
            case.title = sheet.cell(r, 2).value
            case.url = sheet.cell(r, 3).value
            case.method = sheet.cell(r, 5).value
            case.params = sheet.cell(r, 4).value
            case.expected = sheet.cell(r, 6).value
            data_list.append(case)
        wb.close()
        return data_list

    # 方法一：会写测试结果
    def write_result(self, row, column, value):
        '''写回测试结果到Excel中'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, column).value = value
        wb.save(self.file_name)
        wb.close()  # 关闭文件的动作

    # #  方法二
    # def write_result(self, sheet_name, caseid, actual, result):
    #     try:
    #         sheet = self.read_excel[sheet_name]
    #     except Exception as e:
    #         mylog.error(e)
    #         raise e
    #     max_row = sheet.max_row
    #     for i in range(2, max_row + 1):
    #         if caseid == sheet.cell(i, 1).value:
    #             sheet.cell(i, 7).value = actual
    #             sheet.cell(i, 8).value = result
    #             break
    #     self.read_excel.save(self.file_name)
    #     self.read_excel.close()


if __name__ == '__main__':
    wb = DoExcel()
    data_list = wb.get_case('login')
    for case in data_list:
        print(case.caseid)

